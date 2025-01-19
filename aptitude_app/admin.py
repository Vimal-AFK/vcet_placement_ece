from django.contrib import admin
from django.utils.html import format_html
from .models import Batch, Student, Result, QuestionPaper, Question
from django.http import HttpResponse
import csv
import openpyxl
from reportlab.pdfgen import canvas
from django.shortcuts import render, redirect
from django.urls import path
from .forms import StudentImportForm
from datetime import datetime
from django.contrib import messages
from django.db import IntegrityError

# Inline for Students within a Batch
class StudentInline(admin.TabularInline):
    model = Student
    extra = 0  # Do not show extra empty rows
    fields = ('roll_number','name', 'section', 'email','average_percentage')  # Relevant fields to display
    readonly_fields = ('name', 'roll_number', 'section', 'email', 'average_percentage')  # Make fields non-editable
    show_change_link = True  # Enable link to StudentAdmin details

@admin.register(Batch)
class BatchAdmin(admin.ModelAdmin):
    list_display = ('batch_year', 'year', 'total_strength')  # Show basic batch info
    search_fields = ('batch_year', 'year')
    ordering = ('batch_year', 'year')
    list_filter = ('year', 'batch_year')
    inlines = [StudentInline]  # Show students in the batch
    fields = ('batch_year', 'year', 'total_strength')  # Add year field

    def s_no(self, obj):
        return list(obj._meta.model.objects.all()).index(obj) + 1
    s_no.short_description = 'S.No'

# Inline for Results within a Student
class ResultInline(admin.TabularInline):
    model = Result
    extra = 0  # Allows one extra inline entry for adding results quickly
    readonly_fields = ('student', 'test_code', 'date_of_exam', 'total_marks_scored', 'percentage', 'status')  # Make fields non-editable

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ('name', 'roll_number', 'batch', 'section','password','average_percentage')
    readonly_fields = ('password','average_percentage')
    search_fields = ('name', 'roll_number', 'email',)
    list_filter = ('batch', 'section', 'gender')
    ordering = ('batch', 'roll_number')
    inlines = [ResultInline]  # Add Result management directly in Student admin view
    fieldsets = (
        ('Personal Info', {
            'fields': ('name', 'date_of_birth', 'gender', 'email', 'phone_number',)  # Add password field
        }),
        ('Academic Info', {
            'fields': ('batch', 'roll_number', 'section', )
        }),
    )
    exclude = ()  # Remove exclusion of password field
    change_list_template = "admin/student_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-students/', self.admin_site.admin_view(self.import_students), name='import_students'),
            path('download-template/', self.admin_site.admin_view(self.download_template), name='download_template'),
        ]
        return custom_urls + urls

    def import_students(self, request):
        if request.method == 'POST':
            form = StudentImportForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                try:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        batch_year, department, name, roll_number, section, date_of_birth, gender, email, phone_number = row
                        if date_of_birth:
                            date_of_birth = datetime.strptime(date_of_birth, '%d.%m.%Y').date()  # Convert date format
                        batch, created = Batch.objects.get_or_create(batch_year=batch_year, department=department)
                        student, created = Student.objects.update_or_create(
                            email=email,
                            defaults={
                                'batch': batch,
                                'name': name,
                                'roll_number': roll_number,
                                'section': section,
                                'date_of_birth': date_of_birth,
                                'gender': gender,
                                'phone_number': phone_number
                            }
                        )
                    messages.success(request, "Students imported successfully.")
                except IntegrityError as e:
                    messages.error(request, f"Error processing file: {e}")
                except ValueError as e:
                    messages.error(request, f"Error processing file: {e}")
                return redirect('admin:aptitude_app_student_changelist')
        else:
            form = StudentImportForm()
        return render(request, 'admin/import_students.html', {'form': form})

    def download_template(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_template.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['batch_year', 'department', 'name', 'roll_number', 'section', 'date_of_birth', 'gender', 'email', 'phone_number'])
        ws.append([2025, 'ECE', 'John Doe', '12345', 'A', '01.01.2000', 'M', 'john.doe@example.com', '1234567890'])
        wb.save(response)
        return response

    def export_students_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students.csv"'
        writer = csv.writer(response)
        writer.writerow(['S.No', 'Batch Year', 'Department', 'Name', 'Roll Number', 'Section', 'Date of Birth', 'Gender', 'Email', 'Phone Number'])
        for idx, student in enumerate(queryset, start=1):
            writer.writerow([idx, student.batch.batch_year, student.batch.department, student.name, student.roll_number, student.section, student.date_of_birth, student.gender, student.email, student.phone_number])
        return response

    export_students_as_csv.short_description = "Export Selected as CSV"

    def export_students_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['S.No', 'Batch Year', 'Department', 'Name', 'Roll Number', 'Section', 'Date of Birth', 'Gender', 'Email', 'Phone Number'])
        for idx, student in enumerate(queryset, start=1):
            ws.append([idx, student.batch.batch_year, student.batch.department, student.name, student.roll_number, student.section, student.date_of_birth, student.gender, student.email, student.phone_number])
        wb.save(response)
        return response

    export_students_as_excel.short_description = "Export Selected as Excel"

    def export_students_as_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="students.pdf"'
        p = canvas.Canvas(response)
        p.drawString(100, 800, "Students")
        y = 750
        for idx, student in enumerate(queryset, start=1):
            p.drawString(100, y, f"{idx}. {student.batch.batch_year} - {student.batch.department} - {student.name} - {student.roll_number} - {student.section} - {student.date_of_birth} - {student.gender} - {student.email} - {student.phone_number}")
            y -= 20
        p.showPage()
        p.save()
        return response

    export_students_as_pdf.short_description = "Export Selected as PDF"

    actions = [export_students_as_csv, export_students_as_excel, export_students_as_pdf]

    def s_no(self, obj):
        return list(obj._meta.model.objects.all()).index(obj) + 1
    s_no.short_description = 'S.No'

def export_as_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="results.csv"'
    writer = csv.writer(response)
    writer.writerow(['S.No', 'Student', 'Test Code', 'Date of Exam', 'Total Marks Scored', 'Percentage', 'Status', 'Attended'])
    for idx, result in enumerate(queryset, start=1):
        writer.writerow([idx, result.student.name, result.test_code, result.date_of_exam, result.total_marks_scored, result.percentage, result.status, result.attended])
    return response

export_as_csv.short_description = "Export Selected as CSV"

def export_as_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="results.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['S.No', 'Student', 'Test Code', 'Date of Exam', 'Total Marks Scored', 'Percentage', 'Status', 'Attended'])
    for idx, result in enumerate(queryset, start=1):
        ws.append([idx, result.student.name, result.test_code, result.date_of_exam, result.total_marks_scored, result.percentage, result.status, result.attended])
    wb.save(response)
    return response

export_as_excel.short_description = "Export Selected as Excel"

def export_as_pdf(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="results.pdf"'
    p = canvas.Canvas(response)
    p.drawString(100, 800, "Results")
    y = 750
    for idx, result in enumerate(queryset, start=1):
        response['Content-Disposition'] = f'attachment; filename="results{result.test_code}.pdf"'
        p.drawString(100, y, f"{idx}. {result.student.name} - {result.test_code} - {result.date_of_exam} - {result.total_marks_scored} - {result.percentage} - {result.status} - {result.attended}")
        y -= 20
    p.showPage()
    p.save()
    return response

export_as_pdf.short_description = "Export Selected as PDF"

@admin.register(Result)
class ResultAdmin(admin.ModelAdmin):
    list_display = ('student', 'test_code', 'date_of_exam', 'total_marks_scored', 'percentage', 'status', 'attended')
    search_fields = ('student__name', 'test_code')
    list_filter = ('status', 'test_code', 'date_of_exam', 'student__batch', 'student__section', 'student__gender')  # Add filters
    ordering = ('-date_of_exam',)
    autocomplete_fields = ('student',)
    #readonly_fields = ('student', 'test_code', 'date_of_exam', 'total_marks_scored', 'percentage', 'status')  # Make fields non-editable
    actions = [export_as_csv, export_as_excel, export_as_pdf]  # Add export actions

    def s_no(self, obj):
        return list(obj._meta.model.objects.all()).index(obj) + 1
    s_no.short_description = 'S.No'

admin.site.site_header = "VCET Placement Cell of ECE"
admin.site.site_title = "VCET Placement Admin"
admin.site.index_title = "Welcome to Admin Panel"




class QuestionInline(admin.StackedInline):
    model = Question
    extra = 1  # Number of extra blank forms shown
    max_num = None  # Allows adding unlimited questions
    fields = ('question_text', 'question_image', 'mark', 'option_A', 'option_B', 'option_C', 'option_D', 'correct_option')
    readonly_fields = ('image_preview',)

    def image_preview(self, obj):
        if obj.question_image and hasattr(obj.question_image, 'url'):
            return format_html('<img src="{}" style="max-height: 200px; max-width: 200px;" />', obj.question_image.url)
        return "No image"
    image_preview.short_description = 'Image Preview'

@admin.register(QuestionPaper)
class QuestionPaperAdmin(admin.ModelAdmin):
    list_display = ('paper_title', 'paper_code', 'time_limit', 'total_marks', 'is_practice_paper', 'is_assessment_paper', 'preview')
    list_filter = ('is_practice_paper', 'is_assessment_paper')
    search_fields = ('paper_title', 'paper_code', 'paper_description')
    readonly_fields = ('total_marks',)
    inlines = [QuestionInline]

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        obj.calculate_total_marks()  # Automatically calculate total marks after save

    def preview(self, obj):
        return format_html('<a href="{}" target="_blank">Preview</a>', obj.get_absolute_url())
    preview.short_description = 'Preview'

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('question_text', 'mark')
    search_fields = ('question_text',)
    list_filter = ('question_paper__paper_title','question_paper__paper_code')